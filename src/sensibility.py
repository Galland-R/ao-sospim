import sys
from pathlib import Path                                

root = Path(__file__).resolve().parent         
sys.path.append(str(root)) 

from src import dataset_tools as dt
from src import metrics as mt
from src import image_io as iio
import config

import numpy as np
import matplotlib.pyplot as plt
from pathlib import Path
from functools import partial




# Sensibilité 


def coefficient_a(valeurs, x_alpha, plot=True):
    """
    Les valeurs sont déjà normalisé 
    Fait un fit quadratique y = a*x^2 + b*x + c sur les valeurs (déjà normalisées)
    et retourne uniquement le coefficient a.
    """
    y = np.array(valeurs)
    x = x_alpha

    a, b, c = np.polyfit(x, y, 2)

     # Calcul du R²
    y_pred = a * x**2 + b * x + c
    ss_res = np.sum((y - y_pred) ** 2)
    ss_tot = np.sum((y - np.mean(y)) ** 2)
    r2 = 1 - ss_res / ss_tot

    if plot:
        x_fit = np.linspace(x.min(), x.max(), 200)
        y_fit = a * x_fit**2 + b * x_fit + c

        plt.figure(figsize=(7, 5))
        plt.scatter(x, y, color='tab:blue', label='Données normalisées', zorder=3)
        plt.plot(x_fit, y_fit, color='tab:red',
                 label=f'Fit quadratique (R² = {r2:.4f})')
        plt.xlabel('x')
        plt.ylabel('y (normalisé)')
        plt.title('Fit quadratique y = a·x² + b·x + c')
        plt.legend()
        plt.grid(True, alpha=0.3)
        plt.show()
    return a




def coefficient_N(courbes, x_alpha, methode="std", mode="fit", plot=True):
    """
    Calcule le coefficient N (aire entre ymin et ymax).

    Entrée :
    courbes est une liste [[a1,a2, ...], [b1,b2,...], ...]
    alpha est la valeur max d'aberrations induites

    methode : "std" (moyenne ± écart-type) ou "minmax" (vrai min/max observé)
    mode : "direct" (aire sur les points bruts) ou "fit" (aire après fit quadratique)
    plot : si True, affiche y_moy, ymin et ymax
    """
    moy = np.array([np.mean(c) for c in courbes])

    if methode == "std":
        std = np.array([np.std(c) for c in courbes])
        ymin, ymax = moy - std, moy + std
    elif methode == "minmax":
        ymin = np.array([np.min(c) for c in courbes])
        ymax = np.array([np.max(c) for c in courbes])
    else:
        raise ValueError("methode doit être 'std' ou 'minmax'")

    x = x_alpha

    if mode == "direct":
        diff = np.abs(ymax - ymin)
        N = np.trapezoid(diff, x)
    elif mode == "fit":
        coeffs_min = np.polyfit(x, ymin, 2)
        coeffs_max = np.polyfit(x, ymax, 2)

        x_fine = np.linspace(x.min(), x.max(), 500)
        y_min_fit = np.polyval(coeffs_min, x_fine)
        y_max_fit = np.polyval(coeffs_max, x_fine)

        diff = np.abs(y_max_fit - y_min_fit)
        N = np.trapezoid(diff, x_fine)
    else:
        raise ValueError("mode doit être 'direct' ou 'fit'")

    if plot:
        plt.figure(figsize=(7, 5))
        plt.plot(x, moy, color='tab:blue', marker='o', label='y_moy')
        plt.plot(x, ymin, color='tab:green', marker='o', linestyle='--', label='y_min')
        plt.plot(x, ymax, color='tab:orange', marker='o', linestyle='--', label='y_max')
        plt.fill_between(x, ymin, ymax, color='gray', alpha=0.2)

        if mode == "fit":
            plt.plot(x_fine, y_min_fit, color='tab:green', alpha=0.5)
            plt.plot(x_fine, y_max_fit, color='tab:orange', alpha=0.5)

        plt.xlabel('x')
        plt.ylabel('y')
        plt.title(f'Enveloppe ({methode}, {mode}) — N = {N:.4f}')
        plt.legend()
        plt.grid(True, alpha=0.3)
        plt.show()

    return N


# Exportation Excel

from openpyxl import Workbook

def create_analysis_excel(filename):
    """
    Crée un fichier Excel avec les en-têtes de l'analyse.
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Analyse"

    headers = [
        "Métrique",          # A
        "Depth",             # B
        "Qt aberrations",    # C
        "Jeu",               # D
        "Zernike",           # E
        "abs(a_moy)",        # F
        "N",                 # G
        "Conditions méthode",# H
        "Conditions mode"    # I
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    wb.save(filename)


def append_analysis_result(filename, values):
    wb = load_workbook(filename)
    ws = wb["Analyse"]

    ws.append(values)

    wb.save(filename)

# Récupérer les fichiers 


def construire_chemin_dossier(base_dir, profondeur, quantite, jeu, metrique="mean_ROI"):
    """
    Construit le chemin vers le dossier contenant les images.

    Par défaut, on s'arrête au niveau du jeu.
    Exception : si metrique vaut "DAPI" ou "Gallery", on ajoute
    le sous-dossier correspondant.

    Exemples :
    construire_chemin_dossier(base, "Coverslip", "No_aber", 1)
    -> base/Coverslip/No_aber/1

    construire_chemin_dossier(base, "Coverslip", "No_aber", 1, "std")
    -> base/Coverslip/No_aber/1   (inchangé, on s'arrête au jeu)

    construire_chemin_dossier(base, "Coverslip", "No_aber", 1, "DAPI")
    -> base/Coverslip/No_aber/1/Dapi

    construire_chemin_dossier(base, "Coverslip", "No_aber", 1, "Gallery")
    -> base/Coverslip/No_aber/1/Gallery
    """
    chemin = Path(base_dir) / profondeur / quantite / str(jeu)

    if metrique in ("Dapi", "Gallery"):
        chemin = chemin / metrique

    return chemin


def obtenir_images_mode(base_dir, profondeur, quantite, jeu, mode_zernike, metrique="mean_ROI"):
    """
    Récupère les 9 images (référence + 8 niveaux d'aberration)
    d'un mode Zernike donné.
    """
    dossier = construire_chemin_dossier(
        base_dir,
        profondeur,
        quantite,
        jeu,
        metrique
    )

    print(f"Traitement du dossier : {dossier}", flush=True)

    if not dossier.is_dir():
        raise FileNotFoundError(f"Dossier introuvable : {dossier}")

    toutes_images = dt.list_tif_images(dossier)
    chemins_mode = dt.recuperer_images_par_mode(toutes_images, mode_zernike)
    x_alpha = [dt.extraire_quantite_alpha(p) for p in chemins_mode]
    images_mode = [iio.make_2d_image(iio.load_tif(chemin), mode=metrique) for chemin in chemins_mode]

    return x_alpha, images_mode


def obtenir_matrice_metrique(
    base_dir,
    profondeur,
    quantite,
    mode_zernike,
    metrique,
    jeux=(1, 2, 3, 4, 5),
):
    """
    Construit la matrice des valeurs de métrique pour un mode Zernike donné.
    """

    # Choix de la fonction de calcul
    if metrique == "mean_ROI":
        calcul_metrique = mt.imax
    else:
        NA, Lambda, r_min_px, r_max_px = config.obtenir_parametres_metrique(profondeur, metrique)
        calcul_metrique = partial(
            mt.fft_band,
            NA=NA,
            Lambda=Lambda,
            r_min_px=r_min_px,
            r_max_px=r_max_px,
        )

    matrice = []
    x_reference = None

    for jeu in jeux:
        x_jeu, images_mode = obtenir_images_mode(
            base_dir,
            profondeur,
            quantite,
            jeu,
            mode_zernike,
            metrique,
        )

        y_jeu = [calcul_metrique(p) for p in images_mode]

        if x_reference is None:
            x_reference = x_jeu

        matrice.append(y_jeu)

    return x_reference, matrice

def normaliser_par_ligne(matrice):
    """
    Normalise chaque ligne (jeu) par son propre maximum.
    matrice : liste de listes, une ligne par jeu (ex: 5 lignes x 9 colonnes)

    Renvoie une nouvelle matrice de même forme, où chaque ligne
    est divisée par le max de cette ligne.
    """
    matrice_np = np.array(matrice)  # shape (n_jeux, n_points)

    max_par_ligne = matrice_np.max(axis=1, keepdims=True)  # un max par ligne, shape (n_jeux, 1)

    if np.any(max_par_ligne == 0):
        raise ValueError("Une ligne a un maximum égal à 0, normalisation impossible.")

    matrice_normalisee = matrice_np / max_par_ligne  # broadcast automatique sur les lignes

    return matrice_normalisee.tolist()

